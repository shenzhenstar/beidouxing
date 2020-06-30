#coding=utf-8
#date=2020/06/22 0013

# pip install openpyxl
from openpyxl import load_workbook
from ai_work.int_get import IniOps
from ai_work import excel_get
import os
from ai_work.find_file import FindFile
PATH=lambda p: os.path.abspath(
    os.path.join(os.path.dirname(__file__),p)
)


#从配置文件获取Excel文件的格式配置信息
excel_conf = IniOps(PATH(r'../config.ini')).get_excel_conf()
file_conf = IniOps(PATH(r'../config.ini')).get_file_conf()


class WriteExcel(object):
    def __init__(self, excel, sheet_name=None):
        """
        :param excel: Excel文件路径
        :param sheet_name: 要操作的表名
        """
        # 获得工作簿：打开目标表格，再打开目标表单
        self.wb=load_workbook(PATH(excel))
        self.file_path = PATH(excel)
        if sheet_name:
            # 如果知道工作表名，就获取对应的工作表
            self.ws=self.wb[sheet_name]
        else:
            # 如果不知道工作表名，那就获取当前活动的工作表
            self.ws=self.wb.active
    def working(self):

        # 读取
        put_column = excel_conf['put_column']
        put_column_nu = excel_conf['put_column_nu']
        # 左上角的姓名、部门、工号的写入。b代表行号
        for b in range(1,6):
            # 从b表中读取姓名、部门、工号所在行
            b_name = self.ws[str(put_column)+str(b)].value
            b_dept = self.ws[str(put_column)+str(b)].value
            b_deptno = self.ws[str(put_column)+str(b)].value

            # 如果匹配到对应字段，则在该单元格的下一列写入数据
            if "姓名" in str(b_name):
                self.ws.cell(row=b, column=int(put_column_nu)+1).value = excel_conf['name']
            if "部门" in str(b_dept):
                self.ws.cell(row=b, column=int(put_column_nu)+1).value = excel_conf['dept']
            if "工号" in str(b_deptno):
                self.ws.cell(row=b, column=int(put_column_nu)+1).value = excel_conf['wno']


        # 右下角的【员工签字确认】写入,c代表行号
        for c in range(37,50):
            b_ygqz = self.ws[excel_conf['put_weekend_s']+str(c)].value
            if "员工签字" in str(b_ygqz):
                # 12是代表L列，c+1代表在“员工签字确认的下一列签字”
                self.ws.cell(row=c+1, column=int(excel_conf['put_weekend_s_nu'])).value = excel_conf['name']


        # 获取打卡数据
        a_file_name = FindFile(file_conf['a_excel'])
        a_file_string = '../file_data/' + str(a_file_name.find_file())
        data_from = excel_get.ExcelOps(PATH(a_file_string))
        datas_list = data_from.read_data()


        # 遍历打卡数据
        for i in range(1,len(datas_list)+1):
            # 索引从0开始
            datas = datas_list[i-1]
            # print(datas,len(datas))
            # 如果 上班打卡 和 下班打卡 都有一个存在（即有上班），则继续
            """注意：datas的索引是根据获取a表数据的长度而定，如有改动，需变动datas的索引"""
            if datas[1] or datas[2] :
                # print(datas[2],type(datas[2]))
                # 读取 打卡的日期    ---因为datas是元祖，datas[2]的数据是字符串，所以可以对datas[2]进行切片
                if datas[1] :
                    work_date = datas[1][3:5]
                else:
                    work_date = datas[2][3:5]
                # print(work_date)
                # print(work_date[0],type(work_date[0]))
                # 如果打卡的日期是 01 - 09，则识别为 1- 9
                if str(work_date[0]) == '0':
                    # print(work_date[1])
                    daka_date = work_date[1]
                else:
                    # print(work_date)
                    daka_date = work_date


                # 对应日期行的列里写入打卡时间！
                for j in range(7,37):
                    # 获取要写入表的日期
                    b_excel_date = self.ws[str(excel_conf['get_date'])+str(j)].value
                    # print(b_excel_date,type(b_excel_date))
                    # 获取第j行的星期
                    b_excel_week = self.ws[str(excel_conf['get_week'])+str(j)].value
                    # 获取第j行的正常工作时长（单位：h）
                    b_excel_worktime = self.ws[str(excel_conf['put_work_s'])+str(j)].value
                    # 获取第j行的加班工作时长（单位：h）
                    b_excel_other_worktime = self.ws[str(excel_conf['put_weekend_s'])+str(j)].value

                    # 如果从打卡时间里匹配到日期，则在该行写入打卡数据,j是行
                    if str(b_excel_date) == str(daka_date):
                        # 上班打卡时间
                        self.ws.cell(row = j, column = int(excel_conf['put_up_nu'])).value = datas[1]
                        # 下班打卡时间
                        self.ws.cell(row = j, column = int(excel_conf['put_off_nu'])).value = datas[2]

                        # 如果是规定的正常上班，则输出正常上班的日志
                        if str(b_excel_worktime) == '8':
                            # 因为考虑到可能有上班没打卡，或者下班没打卡，所以进行分支输出日志，如果有上班或者下班其中一个没打卡，则输出提示！
                            if not datas[1]:
                                print('正常上班：星期'+str(b_excel_week)+', '+str(b_excel_date)+'号'+'【注意】上班没打卡数据，请注意！')
                            if not datas[2]:
                                print('正常上班：星期'+str(b_excel_week)+', '+str(b_excel_date)+'号'+'【注意】下班没打卡数据，请注意！')
                            else:
                                print('正常上班：星期'+str(b_excel_week)+', '+str(b_excel_date)+'号')
                        else:
                            # 如果不是规定的正常上班日期，则判断是否是周末加班（默认是周末加班，节假日加班需在Excel手动修改）
                            if str(b_excel_week) in '六、日':
                                # 如果是周末加班，则在【周末加班一列】输入工时
                                self.ws.cell(row=j, column=int(excel_conf['put_weekend_s_nu'])).value = 8
                                # 在备注一列说明加班信息
                                self.ws.cell(row=j, column=int(excel_conf['week_bz'])).value = excel_conf['week_bz_info']
                                if not datas[1]:
                                    print('周末加班：星期' + str(b_excel_week) + ', ' + str(b_excel_date) + '号' + '【注意】上班没打卡数据，请注意！')
                                if not datas[2]:
                                    print('周末加班：星期' + str(b_excel_week) + ', ' + str(b_excel_date) + '号' + '【注意】下班没打卡数据，请注意！')
                                else:
                                    print('周末加班：星期' + str(b_excel_week) + ', ' + str(b_excel_date) + '号')
                            else:
                                if not datas[1]:
                                    print('打卡时间是：'+str(b_excel_date)
                                          +'号，请查看是否为节假日加班!，如果是，请手动在节假日一列输入加班时长。' + '【注意】上班没打卡数据，请注意！')
                                if not datas[2]:
                                    print('打卡时间是：'+str(b_excel_date)
                                          +'号，请查看是否为节假日加班!，如果是，请手动在节假日一列输入加班时长。' + '【注意】下班没打卡数据，请注意！')
                                else:
                                    print('打卡时间是：'+str(b_excel_date)+'号，请查看是否为节假日加班!，如果是，请手动在节假日一列输入加班时长。')

        # 保存操作
        self.wb.save(self.file_path)


if __name__ == '__main__':
    print(file_conf)
    file_name = FindFile(file_conf['b_excel'])
    file_string = '../file_data/'+str(file_name.find_file())
    ce = WriteExcel(file_string)
    ret = ce.working()
    # print(1)