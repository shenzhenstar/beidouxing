#coding=utf-8
#date=2020/06/13 0013

# pip install openpyxl
from openpyxl import load_workbook
from ai_work.int_get import IniOps
import os
from ai_work.find_file import FindFile
PATH=lambda p: os.path.abspath(
    os.path.join(os.path.dirname(__file__),p)
)


#从配置文件获取Excel文件的格式配置信息
excel_conf = IniOps(PATH('../config.ini')).get_excel_conf()
file_conf = IniOps(PATH(r'../config.ini')).get_file_conf()

class ExcelOps(object):
    """
    Excel文件操作方法
    """

    def __init__(self, excel, sheet_name=None):
        """
        :param excel: Excel文件路径
        :param sheet_name: 要操作的表名
        """
        # 获得工作簿
        self.wb=load_workbook(PATH(excel))
        if sheet_name:
            # 如果知道工作表名，就获取对应的工作表
            self.ws=self.wb[sheet_name]
        else:
            # 如果不知道工作表名，那就获取当前活动的工作表
            self.ws=self.wb.active
        # 获取工作表的行数
        self.rows=self.ws.max_row
        # 获取工作表的列数
        self.cols=self.ws.max_column
        return
    def read_data(self):
        """
        从Excel文件读取测试用例，并返回
        :param start: 从第几条用例开始读取
        :param count: 读取几条用例，count为0代表读取全部
        :return: 返回打包好的数据
        """
        sheet=self.ws
        # 定义空列表，存储所有的数据
        datas=[]

        """
        get_name  #Bl列--姓名
        get_pno   #C列--身份证号
        get_up    #H列--上班打卡
        get_off   #I列--下班打卡
        """
        # 从配置文件获取姓名、身份证号
        name = excel_conf['name']
        pno = excel_conf['pno']

        # 结束行是所有数据行数最后一行
        end_row_num=self.rows+1

        for row_num in range(2,end_row_num):
            """读取单元格的data"""
            # 读取姓名
            get_name = sheet[excel_conf['get_name']+str(row_num)].value
            # 读取身份证号
            get_pno = sheet[excel_conf['get_pno']+str(row_num)].value
            # print(get_name)
            # print(get_pno)
            # 如果读取的身份证和姓名与输入的姓名和身份证都一致，则获取打卡时间
            if str(get_name) == name and str(get_pno) == pno:
                work_up = sheet[excel_conf['get_up']+str(row_num)].value
                work_off = sheet[excel_conf['get_off']+str(row_num)].value
                # 把读取到的[姓名、身份证号、上班打卡时间、下班打卡时间] 打包一起返回
                datas.append((name, pno, work_up, work_off))
            # print(len(datas))
        return datas

if __name__ == '__main__':
    file_name = FindFile(file_conf['a_excel'])
    file_string = '../file_data/'+str(file_name.find_file())
    ce=ExcelOps(file_string)
    ret=ce.read_data()
    print(ret)
    for i in ret:
        print(i)